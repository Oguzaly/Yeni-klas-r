from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
from openpyxl import Workbook,load_workbook
import pytest

driver = webdriver.Chrome('C:/Users/HP/Desktop/Oğuz/workfile/driver/chromedriver.exe')
driver.get('https://172.27.0.228/atlasui/login')

def enter():
    global url
    gelismis = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"button[id=details-button]"))).click()
    git = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"a[id='proceed-link']"))).click()
    login = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[placeholder='Username']")))
    password = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[placeholder='Password']")))
    login.send_keys('TestUser')
    password.send_keys('TestUser')
    gitlogin = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"button[class='btn btn-primary btn-block']")))
    gitlogin.click()
    loginalready = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,'//button[contains(.,"Login")]'))).click()
    #firstelement = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"span[class='font-weight-semibold font-size-14']")))
    element1=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'ATLAS')]")))
    url=driver.current_url #Giriş yaptıktan sonra gelen dashboard sayfası
enter()
def dealmanagement():
    global url2
    element1=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'ATLAS')]"))).click() #Yuhiiiiii buldummm Thank Selenium Ide
    element2=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'VOD Operations')]"))).click()
    element3=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'Deal Management')]"))).click()
    contractelement = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//a[contains(.,'Contract Management')]")))#Sayfadaki bir element
    url2=driver.current_url
dealmanagement()

def contentmanagement():

    try:
        wp = load_workbook('C:/Users/HP/Desktop/playground/Regresyon/DummyData/content.xlsx')
        ws = wp.active

        for i in range(2,3)    
        contentmanagementtab=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//a[contains(.,'Content Management')]"))).click()
        newcontent=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//button[contains(.,'New Content ')]"))).click()
        contentname=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[contains(.,'Content Name')]"))).send_keys(ws['A{}'.format()])


    except :
        raise

contentmanagement()