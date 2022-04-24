from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
from openpyxl import Workbook,load_workbook
import pytest

driver = webdriver.Chrome('C:/Users/HP/Desktop/Oğuz/workfile/driver/chromedriver.exe')
driver.get('https://172.27.0.228/atlasui/login')

def enter():
    global url
    gelismis = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"button[id=details-button]"))).click()
    git = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"a[id='proceed-link']"))).click()
    login = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[placeholder='Username']")))
    password = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[placeholder='Password']")))
    login.send_keys('TestUser')
    password.send_keys('TestUser')
    gitlogin = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"button[class='btn btn-primary btn-block']")))
    gitlogin.click()
    loginalready = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,'//button[contains(.,"Login")]'))).click()
    #firstelement = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"span[class='font-weight-semibold font-size-14']")))
    element1=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'ATLAS')]")))
    url=driver.current_url #Giriş yaptıktan sonra gelen dashboard sayfası
enter()
def dealmanagement():
    global url2
    element1=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'ATLAS')]"))).click() #Yuhiiiiii buldummm Thank Selenium Ide
    element2=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'VOD Operations')]"))).click()
    element3=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(.,'Deal Management')]"))).click()
    contractelement = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//a[contains(.,'Contract Management')]")))#Sayfadaki bir element
    url2=driver.current_url
dealmanagement()

def licensemanagement():

    global lisansaddmessage
    global lisanssavemessage

    wp = load_workbook('C:/Users/HP/Desktop/playground/Regresyon/DummyData/licensedata.xlsx')
    ws = wp.active
    licenseelement = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//a[contains(.,'License Management')]"))).click()
    editlicense =WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//tr[1]/td[8]/div/div[1]/a/i"))).click()
    licensename = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@name='license.name']")))
    ServiceType = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select Type']")))



    elements=[licensename,ServiceType]
    for j in elements:
        sleep(0.5)
        j.clear()

        # sleep(0.5)
        # ServiceType = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@required placeholder='Select Type']"))).send_keys(Keys.ENTER)
        # sleep(0.5)
        # ServiceType.clear()
        # sleep(0.5)
        # ServiceType.send_keys(ws['B{}'.format(i)].value)
        # LicenseStartDate = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='License Start Date']"))).send_keys(ws['C{}'.format(i)].value,Keys.TAB)
        # LicenseEndDate = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='License End Date']"))).send_keys(ws['D{}'.format(i)].value,Keys.TAB)
        # Contract = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select Contract']")))
        # sleep(0.5)
        # Contract.send_keys('Oğuzhan')
        # sleep(0.5)
        # Contract.send_keys(Keys.DOWN,Keys.ENTER)
        # Platform = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select platform']")))
        # platformtext=['IPTV','Web','Mobile','Smart']
        # for i in platformtext:
        #     Platform.send_keys(i)
        #     sleep(0.5)
        #     Platform.send_keys(Keys.DOWN,Keys.ENTER)
        #     sleep(0.5)
        # Region = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select region']")))
        # sleep(0.5)
        # Region.send_keys('ALL')
        # sleep(0.5)
        # Region.send_keys(Keys.DOWN,Keys.ENTER)
        # sleep(0.5)
        # Currency = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select Currency']")))
        # sleep(0.5)
        # Currency.send_keys('t')
        # sleep(0.5)
        # Currency.send_keys(Keys.DOWN,Keys.ENTER)
        # runType = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//select[@name='runType']")))
        # sleep(0.5)
        # runType.send_keys('u')
        # sleep(0.5)
        # runType.send_keys(Keys.ENTER) #DropDown list de seçerken 2 kez enter yapman gerekebiliyor.
        # sleep(0.5)
        # runType.send_keys(Keys.ENTER)
        # BoxOffice = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Select Box Office']")))
        # sleep(0.5)
        # BoxOffice.send_keys('u')
        # sleep(0.5)
        # BoxOffice.send_keys(Keys.DOWN,Keys.ENTER)
        # Add = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//button[contains(.,' Add ')]"))).click()
        # sleep(0.5)
        #

    sleep(1)
#    editlicense =WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//tr[1]/td[8]/div/div[1]/a/i"))).click()
    sleep(1)

#     deletelicense =  WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//tr[1]/td[8]/div/div[2]/a/i"))).click()
#     sleep(1)
#     delete = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"//button[contains(.,'Delete')]"))).click()
licensemanagement()

class TestClass:
    def test_mainboard(self):
        assert url == 'https://172.27.0.228/atlasui/atlas/cms/dashboard'
    def test_dealmanagement(self):
        assert url2 == 'https://172.27.0.228/atlasui/atlas/cms/dealing/contract'
    def test_LicenseManagement(self):
        assert url3 == 'https://172.27.0.228/atlasui/atlas/cms/dealing/license'
    def test_lisansaddmessage(self):
        assert lisansaddmessage == True
    def test_lisanssavemessage(self):
        assert lisanssavemessage == True
