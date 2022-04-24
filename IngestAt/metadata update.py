import psycopg2
from time import sleep
import requests
from openpyxl import Workbook,load_workbook

#kirala izle içerikler için bekletiyorsun 
wp = load_workbook('C:/Users/HP/Desktop/ingestat/ingestat.xlsx')
ws = wp.active


for row in ws['A']:
    
        
    
    url = "http://10.98.225.178:8090/contents/"+'{}'.format(row.value)+"/ingest?ingestPoster=false&ingestVideo=false"
    payload={}
    headers = {}
    
    print(row.value , 'content id li içeriğe ingest atıldı')
        
        
#    print(url)
#    print(i)

    response = requests.request("POST", url, headers=headers, data=payload)
    print(response.text)
