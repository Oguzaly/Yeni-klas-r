import requests
from openpyxl import Workbook,load_workbook

wp=load_workbook('C:/Users/HP/Desktop/playground/Bulk Epg Ingest/data/data.xlsx')
ws=wp.active

for i in range(1,2):
    url="http://10.98.225.178:8091/channels/"+'{}'.format(ws['A{}'.format(i)].value)+"/programs?releaseDate="+'{}'.format(ws['B{}'.format(i)].value)+""
    payload={}
    headers = {}
    print(url)
    response = requests.request("GET", url, headers=headers, data=payload)
    print(response.text)
    url1 = "http://10.98.225.178:8091/programs/bulk"
    payload1=response
    headers1 = {'Content-Type': 'application/json'}
    response = requests.request("PUT", url1, headers=headers1, data=payload1)
    print(response.text)